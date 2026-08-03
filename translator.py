from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class ConversionError(Exception):
    pass


@dataclass
class ConversionResult:
    excel_path: Path
    pdf_path: Path
    report_path: Path
    package_count: int
    page_count: int
    residual_chinese_count: int


CHINESE_RE = re.compile(r"[\u4e00-\u9fff]+")
BASE_DIR = Path(__file__).resolve().parent
PDF_FONT_REGULAR = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"
PDF_FONTS_READY = False


@dataclass
class ResidualChineseFinding:
    source: str
    location: str
    text: str


@dataclass
class QualityFinding:
    location: str
    text: str


PHRASES = {
    "装     箱     清     单": "PACKING LIST",
    "装箱清单": "PACKING LIST",
    "合同号:\nCONTRACT": "CONTRACT NO.",
    "合同编号": "Contract No.",
    "合同号": "Contract No.",
    "生产编号": "Manufacturing No.",
    "设备编号": "Equipment No.",
    "电梯型号:\nTYPE": "Elevator Model",
    "电梯型号": "Elevator Model",
    "轿厢内尺寸:\nCW×CD×CH": "Car Inside Size\nCW x CD x CH",
    "层/站/门:\nF/S/D": "Floors/Stops/Doors\nF/S/D",
    "开门尺寸:\nCAR ENTRANCE": "Car Entrance Size",
    "备注:\nREMARKS": "Remarks",
    "To：生产部、质检部、物流部、工程部": "To: Production Dept., QC Dept., Logistics Dept., Engineering Dept.",
    "To: 生产部、质检部、Logistics部、工程部": "To: Production Dept., QC Dept., Logistics Dept., Engineering Dept.",
    "装箱员/日期：                                装箱主管/日期：                           装箱检验/日期：": "Packer/Date:                         Packing Supervisor/Date:                         Packing Inspector/Date:",
    "装箱员/Date: 装箱主管/Date: 装箱Inspected by/Date:": "Packer/Date:                         Packing Supervisor/Date:                         Packing Inspector/Date:",
    "陈其贝": "Chen Qibei",
    "加纳": "Ghana",
    "项目编号": "Project No.",
    "箱号": "Package No.",
    "状态": "Packing Status",
    "名称": "Description",
    "序号": "No.",
    "代号": "Code",
    "规格": "Specification",
    "数量": "Qty",
    "单位": "Unit",
    "备注": "Remarks",
    "编制": "Prepared by",
    "审核": "Reviewed by",
    "检验": "Inspected by",
    "物流": "Logistics",
    "日期": "Date",
    "木箱": "Wooden Case",
    "软包": "Soft Package",
    "裸包": "Bare Package",
    "曳引机箱": "Traction Machine Case",
    "曳引机": "Traction Machine",
    "编码器线": "Encoder Cable",
    "控制柜箱": "Control Cabinet Wooden Case",
    "电控箱": "Control Cabinet Case",
    "电控": "Electrical Control System",
    "电气控制": "Electrical Control System",
    "机械部件箱": "Mechanical Parts Case",
    "机械部件": "Mechanical Parts",
    "导轨箱": "Guide Rail Package",
    "导轨": "Guide Rail",
    "对重块箱": "Counterweight Block Package",
    "对重块": "Counterweight Blocks",
    "轿壁箱": "Car Wall Panel Wooden Case",
    "轿壁": "Car Wall Panel",
    "轿顶轿底箱": "Car Top and Car Platform Case",
    "轿顶轿底": "Car Top and Car Platform",
    "门机+层门装置箱": "Door Operator + Landing Door Device Case",
    "门机+层门装置": "Door Operator + Landing Door Device",
    "直梁": "Upright Beam",
    "对重架": "Counterweight Frame",
    "铝合金框架": "Aluminum Alloy Frame",
    "别墅电梯": "Villa Elevator",
    "轿厢导轨": "Car Guide Rail",
    "对重导轨": "Counterweight Guide Rail",
    "导轨接板": "Guide Rail Fishplate",
    "复合对重块": "Composite Counterweight Block",
    "钢板对重块": "Steel Plate Counterweight Block",
    "操纵壁/前壁": "COP Wall / Front Wall Panel",
    "操纵壁": "COP Wall",
    "前壁": "Front Wall Panel",
    "前壁板": "Front Wall Panel",
    "侧中壁": "Side Center Wall Panel",
    "后侧壁": "Rear Side Wall Panel",
    "后中壁": "Rear Center Wall Panel",
    "侧壁": "Side Wall Panel",
    "立柱": "Upright Post",
    "门头": "Door Header",
    "门楣": "Door Header",
    "小门套": "Narrow Door Jamb",
    "小D套": "Narrow Door Jamb",
    "门套": "Door Jamb",
    "门套连接件": "Door Jamb Connector",
    "D套连接件": "Door Jamb Connector",
    "框架玻璃": "Frame Glass",
    "玻璃": "Glass",
    "轿厢拼装螺栓": "Car Assembly Bolts",
    "外六角螺栓": "Hex Head Bolts",
    "轿门": "Car Door",
    "层门": "Landing Door",
    "厅门": "Landing Door",
    "层站召唤盒": "Landing Call Panel",
    "外呼": "Landing Call Panel",
    "轿顶": "Car Top",
    "轿底平台": "Car Platform",
    "轿底": "Car Platform",
    "轿底梁": "Car Bottom Beam",
    "轿门地坎安装组件": "Car Door Sill Installation Assembly",
    "直流风机": "DC Fan",
    "吊顶": "Ceiling",
    "立梁卡板": "Upright Beam Clamp Plate",
    "护脚板及地坎支架": "Toe Guard and Sill Bracket",
    "门机": "Door Operator",
    "层门装置": "Landing Door Device",
    "地坎组件": "Sill Assembly",
    "层门地坎组件": "Landing Door Sill Assembly",
    "扶手": "Handrail",
    "上下围板": "Upper and Lower Skirting Panels",
    "观光玻璃胶垫": "Glass Rubber Pad for Observation Car",
    "玻璃胶及胶枪": "Glass Adhesive and Gun",
    "观光Glass橡胶垫": "Glass Rubber Pad for Observation Car",
    "千里马Glass胶及枪": "Glass Adhesive and Gun (Qianlima Brand)",
    "瞬时式安全钳": "Instantaneous Safety Gear",
    "瞬时安全钳": "Instantaneous Safety Gear",
    "安全钳": "Safety Gear",
    "含提拉机构、导靴板": "with lifting mechanism and guide shoe plate",
    "安全钳拉杆": "Safety Gear Pull Rod",
    "提拉机构": "Safety Gear Longitudinal Pull Rod",
    "安全钳纵向拉杆": "Safety Gear Longitudinal Pull Rod",
    "拉条支架及橡皮圈": "Pull Rod Bracket and Rubber Ring",
    "光幕": "Light Curtain",
    "对重架组件": "Counterweight Frame Assembly",
    "对重导向轮": "Counterweight Deflector Sheave",
    "轿厢反绳轮": "Car Deflector Sheave",
    "轿厢Deflector Sheave": "Car Deflector Sheave",
    "对重导靴": "Counterweight Guide Shoe",
    "钢带轮": "Belt Sheave",
    "对称件": "Symmetrical Part",
    "轿厢上梁": "Car Top Beam",
    "轿厢架上梁": "Car Sling Upper Beam",
    "轿厢架下梁": "Car Sling Lower Beam",
    "轿架上梁": "Car Sling Upper Beam",
    "轿厢反绳轮梁": "Car Diverting Sheave Beam",
    "轿厢返绳轮": "Car Diverting Sheave",
    "轿厢上梁导靴": "Car Upper Beam Guide Shoe",
    "托架导靴": "Bracket Guide Shoe",
    "支架导靴": "Bracket Guide Shoe",
    "轿厢油杯": "Car Oil Cup",
    "对重油杯": "Counterweight Oil Cup",
    "小方油杯": "Small Square Oil Cup",
    "接油盒": "Oil Collector",
    "支撑件组合件": "Support Assembly",
    "限速器": "Overspeed Governor",
    "主机支架梁": "Machine Support Beam",
    "主机支撑梁": "Machine Support Beam",
    "轿厢侧支撑梁": "Car Side Support Beam",
    "轿厢支撑梁": "Car Support Beam",
    "对重侧支撑梁": "Counterweight Side Support Beam",
    "对重支撑梁": "Counterweight Support Beam",
    "侧梁": "Side Beam",
    "轿架下梁": "Car Bottom Beam",
    "轿厢导轨长支架": "Car Guide Rail Long Bracket",
    "轿厢导轨短支架": "Car Guide Rail Short Bracket",
    "对重导轨长支架": "Counterweight Guide Rail Long Bracket",
    "对重导轨短支架": "Counterweight Guide Rail Short Bracket",
    "导轨长支架": "Guide Rail Long Bracket",
    "导轨短支架": "Guide Rail Short Bracket",
    "长支架": "Long Bracket",
    "短支架": "Short Bracket",
    "导轨连接支架": "Guide Rail Connecting Bracket",
    "轿厢导轨支架": "Car Guide Rail Bracket",
    "轿厢导轨支架底码": "Car Guide Rail Bracket Base",
    "对重导轨支架横档": "Counterweight Guide Rail Bracket Cross Member",
    "对重导轨支架支架": "Counterweight Guide Rail Bracket",
    "对重导轨支架": "Counterweight Guide Rail Bracket",
    "对重导轨支架底码": "Counterweight Guide Rail Bracket Base",
    "导轨支架底座": "Guide Rail Bracket Base",
    "导轨支架": "Guide Rail Bracket",
    "导轨连接板": "Guide Rail Fishplate",
    "导靴": "Guide Shoe",
    "上梁导靴": "Upper Beam Guide Shoe",
    "涨紧轮": "Tension Pulley",
    "张紧轮": "Tension Pulley",
    "紧固件": "Fasteners",
    "ULS开关": "ULS Switch",
    "UKS开关": "UKS Switch",
    "撞弓支架": "Cam Bracket",
    "撞弓": "Cam",
    "下梁导靴": "Lower Beam Guide Shoe",
    "缓冲器底座": "Buffer Base",
    "缓冲器": "Buffer",
    "轿厢缓冲器": "Car Buffer",
    "对重缓冲器": "Counterweight Buffer",
    "电焊条": "Welding Rods",
    "轿厢缓冲器座": "Car Buffer Base",
    "对重缓冲器座": "Counterweight Buffer Base",
    "限位开关固定板": "Limit Switch Mounting Plate",
    "端站开关安装支架": "Terminal Switch Mounting Bracket",
    "隔光板安装支架": "Light Shield Mounting Bracket",
    "隔磁板": "Magnetic Shield Plate",
    "楔块绳头组合": "Wedge Rope Socket Assembly",
    "绳头组合": "Wedge Rope Socket Assembly",
    "绳头支撑件": "Rope End Support Bracket",
    "绳头棒": "Rope Rod",
    "绳头杆": "Rope Rod",
    "U型钢丝绳夹头": "U-type Wire Rope Clip",
    "U型限速器钢丝绳夹": "U-type Governor Rope Clip",
    "钢丝绳平行夹": "Wire Rope Parallel Clamp",
    "钢丝绳夹": "Wire Rope Clip",
    "导轨支架调节垫片": "Guide Rail Bracket Adjusting Shim",
    "导轨润滑油": "Guide Rail Lubricating Oil",
    "曳引机钢丝绳": "Traction Rope",
    "曳引绳": "Traction Rope",
    "钢丝绳": "Traction Rope",
    "限速器钢丝绳": "Governor Rope",
    "限速器绳": "Governor Rope",
    "反绳轮": "Deflector Sheave",
    "绳轮": "Sheave",
    "自喷漆": "Spray Paint",
    "光电安装板": "Photoelectric Mounting Plate",
    "光电转接板": "Photoelectric Adapter Plate",
    "轿顶反绳轮固定板": "Car Top Diverting Sheave Fixing Plate",
    "随行电缆夹": "Traveling Cable Clamp",
    "导轨底座": "Guide Rail Base",
    "随机资料": "Technical Documents",
    "控制柜": "Control Cabinet",
    "轿顶检修箱": "Car Top Inspection Box",
    "残疾人操纵箱": "Disabled Control Panel",
    "底坑检修盒(带灯型)": "Pit Inspection Box (with Light)",
    "三方通话": "Three-way Intercom",
    "轿底超载装置": "Car Platform Overload Device",
    "轿厢感应器": "Car Sensor",
    "单光电感应器": "Single Photoelectric Sensor",
    "限位开关": "Limit Switch",
    "终端限位开关": "Final Limit Switch",
    "极限开关": "Final Limit Switch",
    "运行电缆": "Traveling Cable",
    "光电开关": "Photoelectric Sensor",
    "井道线缆": "Hoistway Cables",
    "技术文件": "Technical Documents",
    "爬梯": "Ladder",
}

REPLACEMENTS = [
    ("Contral Cabinet", "Control Cabinet"),
    ("Bare Packing", "Bare Package"),
    ("Soft Packing", "Soft Package"),
    ("含编码器", "with encoder"),
    ("配安装标准件", "with standard mounting parts"),
    ("含小支架", "with small bracket"),
    ("具体见内部明细", "see internal detailed list"),
    ("含提拉机构、导靴板", "with lifting mechanism and guide shoe plate"),
    ("组装发货（含防护罩）", "Shipped Pre-assembled (with protective cover)"),
    ("与轮梁组装发货", "Shipped Pre-assembled with sheave beam"),
    ("组装发货", "Shipped Pre-assembled"),
    ("方形", "square type"),
    ("L型", "L type"),
    ("C型", "C type"),
    ("含压导板，连接螺栓", "with rail clips and connecting bolts"),
    ("连接限位开关", "Connected to Limit Switch "),
    ("连接Limit Switch", "Connected to Limit Switch "),
    ("按楼层站配", "supplied per floor/stop"),
    ("焊接件", "Welded Part"),
    ("绳头棒端头", "Rod End"),
    ("Rope Rod端头", "Rod End"),
    ("有簧", "with spring"),
    ("夹30绳,3槽", "for 30 rope, 3 grooves"),
    ("夹30绳，4槽", "Clamp for 30 mm Rope, 4 Grooves"),
    ("夹30绳,4槽", "Clamp for 30 mm Rope, 4 Grooves"),
    ("柴机油", "diesel engine oil"),
    ("颜色：", "Color: "),
    ("颜色:", "Color: "),
    ("黑砂", "Black Sand Finish"),
    ("含底座", "with base"),
    ("带紧固件", "With Fasteners"),
    ("带Fasteners", "With Fasteners"),
    ("含地板及减震胶垫", "with floor and vibration damping rubber pads"),
    ("含地坎托架及护脚板", "with sill bracket and toe guard"),
    ("含灯具", "with light fixtures"),
    ("与直梁装配好", "Pre-assembled with the Upright Beam"),
    ("与Upright Beam装配好", "Pre-assembled with the Upright Beam"),
    ("含弹簧及紧固件", "Including Springs and Fasteners"),
    ("含弹簧及Fasteners", "Including Springs and Fasteners"),
    ("含防尘罩，油杯座", "with dust cover and oil cup seat"),
    ("与对重架组装发货", "Shipped Pre-assembled with counterweight frame"),
    ("与Counterweight FrameShipped Pre-assembled", "Pre-assembled with the Counterweight Frame"),
    ("直径", "diameter "),
    ("导轨顶面宽", "Guide Rail Head Width"),
    ("Guide Rail顶面宽", "Guide Rail Head Width"),
    ("一段有连接机器底座孔", "one section has holes for machine base connection"),
    ("米/根", "m/pc"),
    ("升", " L"),
    ("框架Glass", "Frame Glass"),
    ("后Side Wall Panel", "Rear Side Wall Panel"),
    ("Side Wall Panel（后侧）", "Rear Side Wall Panel"),
    ("Side Wall Panel(后侧)", "Rear Side Wall Panel"),
    ("色", "Color"),
]

UNITS = {
    "台": "unit",
    "个": "pcs",
    "件": "pcs",
    "套": "set",
    "组": "set",
    "条": "pcs",
    "根": "pcs",
    "块": "PCS",
    "包": "pack",
    "米": "m",
    "壶": "can",
    "瓶": "bottle",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def has_chinese(value: Any) -> bool:
    return bool(CHINESE_RE.search(clean_text(value)))


def compact_text(value: Any, limit: int = 120) -> str:
    text = re.sub(r"\s+", " ", clean_text(value))
    return text if len(text) <= limit else f"{text[:limit - 3]}..."


def translate(value: Any) -> str:
    source = clean_text(value)
    if not source:
        return ""
    if source in PHRASES:
        return PHRASES[source]
    if source in UNITS:
        return UNITS[source]

    result = source
    for chinese, english in sorted(PHRASES.items(), key=lambda item: len(item[0]), reverse=True):
        result = result.replace(chinese, english)
    for chinese, english in REPLACEMENTS:
        result = result.replace(chinese, english)
    result = re.sub(r"(\d+)层(\d+)站(\d+)门", r"\1F/\2S/\3D", result)
    result = re.sub(r"(\d+(?:\.\d+)?)\s*kg/块", r"\1 kg/PCS", result, flags=re.IGNORECASE)
    result = re.sub(r"P\+Q\s*=\s*(\d+(?:\.\d+)?)\s*kg", r"P+Q = \1 kg", result, flags=re.IGNORECASE)
    result = re.sub(r"V\s*=\s*(\d+(?:\.\d+)?)\s*m/s", r"V = \1 m/s", result, flags=re.IGNORECASE)
    result = re.sub(r"Guide Rail Head Width\s*=\s*(\d+(?:\.\d+)?)", r"Guide Rail Head Width = \1 mm", result)
    result = result.replace("层", "F").replace("站", "S").replace("门", "D")
    result = result.replace("，", ", ").replace("、", ", ").replace("：", ": ").replace("φ", "dia. ")
    return re.sub(r"\s+", " ", result).strip()


def find_packing_sheet(workbook):
    if "装箱单" in workbook.sheetnames:
        return workbook["装箱单"]
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 40), values_only=True):
            row_text = " ".join(clean_text(value) for value in row if value is not None)
            if "装箱清单" in row_text or "装 箱 清 单" in row_text:
                return sheet
    raise ConversionError("没有找到装箱单工作表。请确认文件里有“装箱单”或装箱清单页面。")


def find_detail_starts(sheet) -> list[int]:
    starts = []
    for row in range(1, sheet.max_row + 1):
        value = clean_text(sheet.cell(row=row, column=1).value)
        if "装" in value and "箱" in value and "清" in value and "单" in value:
            if row > 20:
                starts.append(row)
    if starts:
        return starts
    for row in range(1, sheet.max_row + 1):
        labels = [clean_text(sheet.cell(row=row, column=col).value) for col in range(1, 10)]
        if "合同编号" in labels and "电梯型号" in labels:
            starts.append(row - 1)
    if not starts:
        raise ConversionError("没有识别到每个箱号的明细页。请确认装箱单格式和样表相近。")
    return starts


def cell_value(sheet, address: str) -> Any:
    return sheet[address].value


def paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    escaped = translate(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return Paragraph(escaped.replace("\n", "<br/>"), style)


def register_pdf_fonts() -> tuple[str, str]:
    global PDF_FONTS_READY, PDF_FONT_REGULAR, PDF_FONT_BOLD
    if PDF_FONTS_READY:
        return PDF_FONT_REGULAR, PDF_FONT_BOLD

    bundled_regular = BASE_DIR / "assets" / "fonts" / "NotoSansSC-Regular.ttf"
    bundled_bold = BASE_DIR / "assets" / "fonts" / "NotoSansSC-Bold.ttf"
    local_cjk_candidates = [
        Path("/System/Library/Fonts/STHeiti Light.ttc"),
        Path("/System/Library/Fonts/Hiragino Sans GB.ttc"),
    ]

    try:
        if bundled_regular.exists() and bundled_regular.stat().st_size > 1024 * 1024:
            pdfmetrics.registerFont(TTFont("NotoSansSC", str(bundled_regular)))
            PDF_FONT_REGULAR = "NotoSansSC"
            if bundled_bold.exists() and bundled_bold.stat().st_size > 1024 * 1024:
                pdfmetrics.registerFont(TTFont("NotoSansSC-Bold", str(bundled_bold)))
                PDF_FONT_BOLD = "NotoSansSC-Bold"
            else:
                PDF_FONT_BOLD = PDF_FONT_REGULAR
        else:
            for font_path in local_cjk_candidates:
                if not font_path.exists():
                    continue
                try:
                    pdfmetrics.registerFont(TTFont("LocalCJK", str(font_path), subfontIndex=0))
                    PDF_FONT_REGULAR = "LocalCJK"
                    PDF_FONT_BOLD = "LocalCJK"
                    break
                except Exception:
                    continue
            else:
                pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
                PDF_FONT_REGULAR = "STSong-Light"
                PDF_FONT_BOLD = "STSong-Light"
    except Exception:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        PDF_FONT_REGULAR = "STSong-Light"
        PDF_FONT_BOLD = "STSong-Light"

    PDF_FONTS_READY = True
    return PDF_FONT_REGULAR, PDF_FONT_BOLD


def make_styles():
    regular_font, bold_font = register_pdf_fonts()
    sample = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("title", parent=sample["Title"], fontName=bold_font, fontSize=18, leading=22, alignment=TA_CENTER),
        "small": ParagraphStyle("small", parent=sample["Normal"], fontName=regular_font, fontSize=8.5, leading=10, alignment=TA_LEFT),
        "cell": ParagraphStyle("cell", parent=sample["Normal"], fontName=regular_font, fontSize=7.3, leading=8.8, alignment=TA_LEFT),
        "center": ParagraphStyle("center", parent=sample["Normal"], fontName=regular_font, fontSize=7.3, leading=8.8, alignment=TA_CENTER),
        "head": ParagraphStyle("head", parent=sample["Normal"], fontName=bold_font, fontSize=7.2, leading=8.5, alignment=TA_CENTER),
    }


def translate_excel(input_path: Path, output_path: Path) -> None:
    shutil.copy2(input_path, output_path)
    workbook = load_workbook(output_path)
    sheet = find_packing_sheet(workbook)
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and not cell.value.startswith("="):
                cell.value = translate(cell.value)
                cell.alignment = cell.alignment.copy(wrap_text=True, shrink_to_fit=True)
    for column, width in {"C": 24, "D": 24, "H": 28, "I": 18}.items():
        if sheet.column_dimensions[column].width is None or sheet.column_dimensions[column].width < width:
            sheet.column_dimensions[column].width = width
    workbook.save(output_path)


def scan_excel_for_residual_chinese(path: Path) -> list[ResidualChineseFinding]:
    findings: list[ResidualChineseFinding] = []
    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet = find_packing_sheet(workbook)
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and has_chinese(cell.value):
                findings.append(
                    ResidualChineseFinding(
                        source="Excel",
                        location=f"{sheet.title}!{cell.coordinate}",
                        text=compact_text(cell.value),
                    )
                )
    workbook.close()
    return findings


def scan_pdf_for_residual_chinese(path: Path) -> list[ResidualChineseFinding]:
    findings: list[ResidualChineseFinding] = []
    reader = PdfReader(path)
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        for match in CHINESE_RE.finditer(text):
            start = max(match.start() - 35, 0)
            end = min(match.end() + 35, len(text))
            findings.append(
                ResidualChineseFinding(
                    source="PDF",
                    location=f"page {index}",
                    text=compact_text(text[start:end]),
                )
            )
    return findings


def scan_pdf_source_for_residual_chinese(input_path: Path) -> list[ResidualChineseFinding]:
    findings: list[ResidualChineseFinding] = []
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    sheet = find_packing_sheet(workbook)
    starts = find_detail_starts(sheet)
    for index, start in enumerate(starts):
        next_start = starts[index + 1] if index + 1 < len(starts) else None
        meta_cells = {
            f"B{start + 1}": sheet.cell(start + 1, 2).value,
            f"G{start + 1}": sheet.cell(start + 1, 7).value,
            f"G{start + 2}": sheet.cell(start + 2, 7).value,
            f"I{start + 2}": sheet.cell(start + 2, 9).value,
            f"B{start + 3}": sheet.cell(start + 3, 2).value,
            f"G{start + 3}": sheet.cell(start + 3, 7).value,
        }
        for coordinate, value in meta_cells.items():
            translated = translate(value)
            if has_chinese(translated):
                findings.append(
                    ResidualChineseFinding(
                        source="PDF source",
                        location=f"{sheet.title}!{coordinate}",
                        text=compact_text(translated),
                    )
                )

        end = (next_start - 1) if next_start else sheet.max_row
        for row in range(start + 5, min(end, start + 36) + 1):
            for column in [1, 2, 3, 4, 6, 7, 8]:
                value = sheet.cell(row, column).value
                if not clean_text(value):
                    continue
                translated = translate(value)
                if has_chinese(translated):
                    findings.append(
                        ResidualChineseFinding(
                            source="PDF source",
                            location=f"{sheet.title}!{sheet.cell(row, column).coordinate}",
                            text=compact_text(translated),
                        )
                    )
    workbook.close()
    return findings


def parse_item_no(value: Any) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    match = re.fullmatch(r"\d+(?:\.0)?", text)
    return int(float(text)) if match else None


def scan_source_for_quality_findings(input_path: Path) -> list[QualityFinding]:
    findings: list[QualityFinding] = []
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    sheet = find_packing_sheet(workbook)
    starts = find_detail_starts(sheet)
    for index, start in enumerate(starts):
        next_start = starts[index + 1] if index + 1 < len(starts) else None
        end = (next_start - 1) if next_start else sheet.max_row
        expected_no: int | None = None
        for row in range(start + 5, min(end, start + 36) + 1):
            no_value = sheet.cell(row, 1).value
            description = clean_text(sheet.cell(row, 3).value)
            qty = clean_text(sheet.cell(row, 6).value)
            unit = clean_text(sheet.cell(row, 7).value)
            row_text = " ".join(clean_text(sheet.cell(row, column).value) for column in [1, 2, 3, 4, 6, 7, 8])
            if not row_text:
                continue
            if row_text.startswith("装箱员") or description == "备注":
                continue

            item_no = parse_item_no(no_value)
            if item_no is not None:
                if expected_no is not None and item_no != expected_no:
                    findings.append(
                        QualityFinding(
                            location=f"{sheet.title}!A{row}",
                            text=f"No. jumps from {expected_no - 1} to {item_no}; please check missing or duplicated item numbers.",
                        )
                    )
                expected_no = item_no + 1

            if description and qty and not unit:
                translated_description = translate(description)
                findings.append(
                    QualityFinding(
                        location=f"{sheet.title}!G{row}",
                        text=f"Unit is missing for '{translated_description}' with Qty {qty}.",
                    )
                )
    workbook.close()
    return findings


def write_translation_report(
    path: Path,
    findings: list[ResidualChineseFinding],
    quality_findings: list[QualityFinding] | None = None,
) -> None:
    quality_findings = quality_findings or []
    lines = [
        "Translation residual Chinese check",
        "==================================",
        "",
    ]
    if not findings:
        lines.extend([
            "PASS: No Chinese text was found in the generated PDF, the PDF source data, or the translated packing list sheet.",
            "",
        ])
    else:
        lines.extend([
            f"WARNING: {len(findings)} Chinese text fragment(s) were found after translation.",
            "Please add the missing elevator terms to the dictionary and regenerate the files.",
            "",
        ])
        for index, finding in enumerate(findings, start=1):
            lines.append(f"{index}. [{finding.source}] {finding.location}: {finding.text}")
    lines.extend([
        "",
        "Packing list data quality check",
        "===============================",
        "",
    ])
    if not quality_findings:
        lines.extend([
            "PASS: Item numbers are continuous and required Qty/Unit fields look complete.",
            "",
        ])
    else:
        lines.extend([
            f"WARNING: {len(quality_findings)} data quality issue(s) were found.",
            "Please review the source packing list before printing or shipping.",
            "",
        ])
        for index, finding in enumerate(quality_findings, start=1):
            lines.append(f"{index}. {finding.location}: {finding.text}")
    path.write_text("\n".join(lines), encoding="utf-8")


def package_meta(sheet, start: int) -> dict[str, Any]:
    return {
        "contract": sheet.cell(start + 1, 2).value,
        "model": sheet.cell(start + 1, 7).value,
        "package_no": sheet.cell(start + 2, 7).value,
        "status": sheet.cell(start + 2, 9).value,
        "manufacturing_no": sheet.cell(start + 3, 2).value,
        "description": sheet.cell(start + 3, 7).value,
    }


def item_rows(sheet, start: int, next_start: int | None, styles: dict[str, ParagraphStyle]) -> list[list[Paragraph]]:
    rows = [[
        paragraph("No.", styles["head"]),
        paragraph("Code", styles["head"]),
        paragraph("Description", styles["head"]),
        paragraph("Specification", styles["head"]),
        paragraph("Qty", styles["head"]),
        paragraph("Unit", styles["head"]),
        paragraph("Remarks", styles["head"]),
    ]]
    end = (next_start - 1) if next_start else sheet.max_row
    for row in range(start + 5, min(end, start + 36) + 1):
        values = [
            sheet.cell(row, 1).value,
            sheet.cell(row, 2).value,
            sheet.cell(row, 3).value,
            sheet.cell(row, 4).value,
            sheet.cell(row, 6).value,
            sheet.cell(row, 7).value,
            sheet.cell(row, 8).value,
        ]
        if not any(clean_text(value) for value in values):
            continue
        if clean_text(values[0]).startswith("装箱员") or clean_text(values[2]) == "备注":
            continue
        rows.append([
            paragraph(values[0], styles["center"]),
            paragraph(values[1], styles["cell"]),
            paragraph(values[2], styles["cell"]),
            paragraph(values[3], styles["cell"]),
            paragraph(values[4], styles["center"]),
            paragraph(values[5], styles["center"]),
            paragraph(values[6], styles["cell"]),
        ])
    return rows


def add_table_style(table: Table, header: bool = False) -> None:
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.55, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    if header:
        commands.append(("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke))
    table.setStyle(TableStyle(commands))


def extract_logo(sheet, output_dir: Path) -> Path | None:
    images = getattr(sheet, "_images", [])
    if not images:
        return None
    data = images[0]._data()
    suffix = ".png" if data.startswith(b"\x89PNG") else ".jpg"
    logo_path = output_dir / f"logo{suffix}"
    logo_path.write_bytes(data)
    return logo_path


def logo_flowable(logo_path: Path | None):
    if not logo_path:
        return None
    return Image(str(logo_path), width=52 * mm, height=28 * mm)


def build_pdf(input_path: Path, output_path: Path) -> int:
    workbook = load_workbook(input_path, data_only=True)
    sheet = find_packing_sheet(workbook)
    starts = find_detail_starts(sheet)
    logo_path = extract_logo(sheet, output_path.parent)
    styles = make_styles()

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=9 * mm,
        leftMargin=9 * mm,
        topMargin=10 * mm,
        bottomMargin=9 * mm,
    )
    story = []
    cover_logo = logo_flowable(logo_path)
    if cover_logo:
        story.append(cover_logo)
    story.append(Paragraph("PACKING LIST", styles["title"]))

    cover_rows = [[paragraph("Package No.", styles["head"]), paragraph("Description", styles["head"]), paragraph("Packing Status", styles["head"])]]
    for start in starts:
        meta = package_meta(sheet, start)
        cover_rows.append([
            paragraph(meta["package_no"], styles["center"]),
            paragraph(meta["description"], styles["cell"]),
            paragraph(meta["status"], styles["center"]),
        ])
    cover = Table(cover_rows, colWidths=[35 * mm, 100 * mm, 45 * mm], hAlign="CENTER")
    add_table_style(cover, header=True)
    story.extend([Spacer(1, 8 * mm), cover])

    for index, start in enumerate(starts):
        next_start = starts[index + 1] if index + 1 < len(starts) else None
        meta = package_meta(sheet, start)
        story.append(PageBreak())
        page_logo = logo_flowable(logo_path)
        if page_logo:
            story.append(page_logo)
        story.append(Paragraph("PACKING LIST", styles["title"]))
        meta_table = Table(
            [
                [paragraph("Contract No.", styles["head"]), paragraph(meta["contract"], styles["cell"]),
                 paragraph("Elevator Model", styles["head"]), paragraph(meta["model"], styles["cell"])],
                [paragraph("Package No.", styles["head"]), paragraph(meta["package_no"], styles["cell"]),
                 paragraph("Packing Status", styles["head"]), paragraph(meta["status"], styles["cell"])],
                [paragraph("Manufacturing No.", styles["head"]), paragraph(meta["manufacturing_no"], styles["cell"]),
                 paragraph("Description", styles["head"]), paragraph(meta["description"], styles["cell"])],
            ],
            colWidths=[34 * mm, 58 * mm, 34 * mm, 64 * mm],
        )
        add_table_style(meta_table)
        item_table = Table(
            item_rows(sheet, start, next_start, styles),
            colWidths=[12 * mm, 25 * mm, 42 * mm, 40 * mm, 15 * mm, 15 * mm, 41 * mm],
            repeatRows=1,
        )
        add_table_style(item_table, header=True)
        story.extend([
            meta_table,
            Spacer(1, 3 * mm),
            item_table,
        ])

    doc.build(story)
    return len(starts)


def convert_workbook(input_path: Path, output_dir: Path) -> ConversionResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / "packing-list-English.xlsx"
    pdf_path = output_dir / "packing-list-English.pdf"
    report_path = output_dir / "translation-check-report.txt"

    try:
        translate_excel(input_path, excel_path)
        package_count = build_pdf(input_path, pdf_path)
        findings = scan_excel_for_residual_chinese(excel_path)
        findings.extend(scan_pdf_source_for_residual_chinese(input_path))
        findings.extend(scan_pdf_for_residual_chinese(pdf_path))
        quality_findings = scan_source_for_quality_findings(input_path)
        write_translation_report(report_path, findings, quality_findings)
    except Exception as exc:
        if isinstance(exc, ConversionError):
            raise
        raise ConversionError(f"转换失败：{exc}") from exc

    return ConversionResult(
        excel_path=excel_path,
        pdf_path=pdf_path,
        report_path=report_path,
        package_count=package_count,
        page_count=package_count + 1,
        residual_chinese_count=len(findings),
    )
